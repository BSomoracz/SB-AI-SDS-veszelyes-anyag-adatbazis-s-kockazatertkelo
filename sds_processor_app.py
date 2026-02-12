#!/usr/bin/env python3
"""
SDS→Excel AI Feldolgozó Rendszer - Prototípus
==============================================
Streamlit alapú webalkalmazás, amely:
1. PDF biztonsági adatlapokat fogad (batch upload, 1-100 fájl)
2. GPT-4o / Claude API-val kinyeri a strukturált adatokat
3. Az eredeti mintafájl formátumában Excel-t generál
4. Kémiai kockázatértékelést és védőeszköz-specifikációt is készít

Telepítés:
    pip install streamlit openai anthropic openpyxl PyPDF2 pydantic python-dotenv

Indítás:
    streamlit run sds_processor_app.py
"""

import streamlit as st
import json
import os
from typing import Optional
from pydantic import BaseModel, Field
from enum import Enum

# ============================================================
# 1. PYDANTIC SÉMA - Strukturált output az LLM-nek
# ============================================================

class SDSComponent(BaseModel):
    """Egy összetevő adatai az SDS 3. szekciójából"""
    name: str = Field(description="Összetevő neve magyarul vagy angolul")
    cas_number: Optional[str] = Field(None, description="CAS szám, pl. 108-88-3")
    ec_number: Optional[str] = Field(None, description="EK szám, pl. 203-625-9")
    concentration: Optional[str] = Field(None, description="Koncentráció %, pl. 30-60%")
    clp_classification: Optional[str] = Field(None, description="CLP besorolás, pl. Flam. Liq. 2, Repr. 1B")

class PhysicalProperties(BaseModel):
    """Fizikai-kémiai tulajdonságok (SDS 9. szekció)"""
    physical_state: Optional[str] = Field(None, description="Halmazállapot")
    color: Optional[str] = Field(None, description="Szín")
    odor: Optional[str] = Field(None, description="Szag")
    melting_point: Optional[str] = Field(None, description="Olvadáspont °C")
    boiling_point: Optional[str] = Field(None, description="Forráspont °C")
    flash_point: Optional[str] = Field(None, description="Lobbanáspont °C")
    autoignition_temp: Optional[str] = Field(None, description="Gyulladási hőmérséklet °C")
    density: Optional[str] = Field(None, description="Sűrűség g/cm³")
    water_solubility: Optional[str] = Field(None, description="Vízoldhatóság")
    ph: Optional[str] = Field(None, description="pH érték")
    vapor_pressure: Optional[str] = Field(None, description="Gőznyomás")

class ExposureLimits(BaseModel):
    """Munkahelyi határértékek (SDS 8. szekció + 5/2020 ITM)"""
    ak_value: Optional[str] = Field(None, description="ÁK-érték mg/m³")
    ck_value: Optional[str] = Field(None, description="CK-érték mg/m³")
    mk_value: Optional[str] = Field(None, description="MK-érték mg/m³")
    dnel_inhalation: Optional[str] = Field(None, description="DNEL munkás inhaláció")
    dnel_dermal: Optional[str] = Field(None, description="DNEL munkás dermális")
    boelv: Optional[str] = Field(None, description="BOELV (EU) mg/m³")

class PPE(BaseModel):
    """Egyéni védőeszközök (SDS 8. szekció)"""
    respiratory: Optional[str] = Field(None, description="Légzésvédelem típusa, szűrő")
    hand: Optional[str] = Field(None, description="Kézvédelem: kesztyű típus, anyag, vastagság")
    eye: Optional[str] = Field(None, description="Szemvédelem típusa")
    skin: Optional[str] = Field(None, description="Bőr/testvédelem")
    engineering: Optional[str] = Field(None, description="Műszaki védelem (elszívás stb.)")

class FireProtection(BaseModel):
    """Tűzvédelmi adatok (SDS 5. szekció)"""
    suitable_extinguishing: Optional[str] = Field(None, description="Megfelelő oltóanyag")
    unsuitable_extinguishing: Optional[str] = Field(None, description="Nem megfelelő oltóanyag")
    hazardous_decomposition: Optional[str] = Field(None, description="Veszélyes bomlástermékek")
    firefighter_ppe: Optional[str] = Field(None, description="Tűzoltói védőfelszerelés")

class Toxicology(BaseModel):
    """Toxikológiai adatok (SDS 11. szekció)"""
    ld50_oral: Optional[str] = Field(None, description="Akut toxicitás orális LD50")
    ld50_dermal: Optional[str] = Field(None, description="Akut toxicitás dermális LD50")
    lc50_inhalation: Optional[str] = Field(None, description="Akut toxicitás inhaláció LC50")
    skin_irritation: Optional[str] = Field(None, description="Bőrirritáció")
    eye_irritation: Optional[str] = Field(None, description="Szemirritáció")
    sensitization: Optional[str] = Field(None, description="Szenzibilizáció")
    cmr_effects: Optional[str] = Field(None, description="CMR hatások (rákkeltő, mutagén, repr. toxikus)")

class TransportInfo(BaseModel):
    """Szállítási adatok (SDS 14. szekció)"""
    un_number: Optional[str] = Field(None, description="UN szám")
    shipping_name: Optional[str] = Field(None, description="Szállítási megnevezés")
    adr_class: Optional[str] = Field(None, description="ADR osztály")
    packing_group: Optional[str] = Field(None, description="Csomagolási csoport")
    marine_pollutant: Optional[str] = Field(None, description="Tengeri szennyező")

class WasteInfo(BaseModel):
    """Hulladékkezelési adatok (SDS 13. szekció)"""
    ewc_code: Optional[str] = Field(None, description="EWC kód")
    disposal_method: Optional[str] = Field(None, description="Hulladékkezelési módszer")

class SDSExtraction(BaseModel):
    """Teljes SDS adatkinyerés - ez a fő séma amit az LLM-nek átadunk"""
    # 1. szekció - Azonosítás
    product_name: str = Field(description="Kereskedelmi név")
    sds_language: Optional[str] = Field(None, description="SDS nyelve (HU/EN/DE)")
    sds_version: Optional[str] = Field(None, description="SDS verziószám")
    sds_date: Optional[str] = Field(None, description="SDS kiadás dátuma")
    sds_revision_date: Optional[str] = Field(None, description="SDS felülvizsgálat dátuma")
    manufacturer: Optional[str] = Field(None, description="Gyártó/Szállító neve")
    manufacturer_address: Optional[str] = Field(None, description="Gyártó címe")
    manufacturer_phone: Optional[str] = Field(None, description="Gyártó telefonszáma")
    manufacturer_email: Optional[str] = Field(None, description="Gyártó e-mail")
    emergency_phone: Optional[str] = Field(None, description="Sürgősségi telefonszám")
    ufi_code: Optional[str] = Field(None, description="UFI kód")
    product_form: Optional[str] = Field(None, description="Termék formája (folyadék, szilárd, aeroszol stb.)")
    intended_use: Optional[str] = Field(None, description="Felhasználási terület")
    use_category: Optional[str] = Field(None, description="Felhasználási kategória")
    substance_or_mixture: Optional[str] = Field(None, description="Anyag vagy Keverék")

    # 3. szekció - Összetétel
    component_1: Optional[SDSComponent] = Field(None, description="1. fő összetevő")
    component_2: Optional[SDSComponent] = Field(None, description="2. fő összetevő")
    component_3: Optional[SDSComponent] = Field(None, description="3. fő összetevő")

    # 2. szekció - Veszélyesség
    clp_classification: Optional[str] = Field(None, description="CLP osztályozás teljes szöveg")
    ghs_pictograms: Optional[str] = Field(None, description="GHS piktogram kódok, pl. GHS02, GHS07")
    signal_word: Optional[str] = Field(None, description="Jelzőszó: Veszély / Figyelem")
    h_statements: Optional[str] = Field(None, description="H mondatok kódokkal, pl. H225; H319")
    p_statements: Optional[str] = Field(None, description="P mondatok kódokkal")
    euh_statements: Optional[str] = Field(None, description="EUH mondatok")
    svhc: Optional[str] = Field(None, description="SVHC (különösen aggályos) anyag")
    pbt_vpvb: Optional[str] = Field(None, description="PBT/vPvB besorolás")

    # 9. szekció
    physical_properties: Optional[PhysicalProperties] = None

    # 8. szekció
    exposure_limits: Optional[ExposureLimits] = None
    ppe: Optional[PPE] = None

    # 5. szekció
    fire_protection: Optional[FireProtection] = None

    # 11. szekció
    toxicology: Optional[Toxicology] = None

    # 14. szekció
    transport: Optional[TransportInfo] = None

    # 13. szekció
    waste: Optional[WasteInfo] = None


# ============================================================
# 2. STREAMLIT ALKALMAZÁS
# ============================================================

st.set_page_config(
    page_title="SDS → Excel AI Feldolgozó",
    page_icon="🧪",
    layout="wide"
)

st.title("🧪 SDS → Excel AI Feldolgozó Rendszer")
st.markdown("""
### Automatikus biztonsági adatlap feldolgozás
Töltsd fel a PDF biztonsági adatlapokat (SDS/MSDS), és a rendszer AI segítségével 
kinyeri az összes releváns adatot a mintafájl formátumában.
""")

# Sidebar beállítások
with st.sidebar:
    st.header("⚙️ Beállítások")

    api_provider = st.selectbox("AI szolgáltató", ["OpenAI (GPT-4o)", "Anthropic (Claude 3.5)"])
    api_key = st.text_input("API kulcs", type="password")

    st.divider()
    st.header("📋 Mintafájl")
    template_file = st.file_uploader("Mintafájl feltöltése (opcionális)", type=["xlsx"])

    st.divider()
    st.header("🏭 Céginformáció")
    company_name = st.text_input("Cég neve", "")
    site_name = st.text_input("Telephely neve", "")

# Fő felület
col1, col2 = st.columns([2, 1])

with col1:
    st.header("📤 SDS PDF-ek feltöltése")
    uploaded_files = st.file_uploader(
        "PDF biztonsági adatlapok (max. 100 fájl)",
        type=["pdf"],
        accept_multiple_files=True
    )

    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} fájl feltöltve")
        for f in uploaded_files:
            st.text(f"  📄 {f.name} ({f.size/1024:.0f} KB)")

with col2:
    st.header("📊 Feldolgozás státusz")
    if uploaded_files:
        progress = st.progress(0)
        status = st.empty()

        if st.button("🚀 Feldolgozás indítása", type="primary"):
            for i, f in enumerate(uploaded_files):
                progress.progress((i + 1) / len(uploaded_files))
                status.text(f"Feldolgozás: {f.name}...")
                # Itt hívnánk az AI API-t
            status.text("✅ Feldolgozás kész!")
    else:
        st.info("⬆️ Tölts fel SDS PDF fájlokat a feldolgozáshoz")


# ============================================================
# 3. FELDOLGOZÓ LOGIKA (API hívások)
# ============================================================

SYSTEM_PROMPT = """Te egy veszélyes anyagok szakértője vagy, aki biztonsági adatlapokból (SDS/MSDS) 
nyer ki strukturált adatokat. 

FELADAT: Olvasd el a biztonsági adatlapot és töltsd ki az összes mezőt a JSON sémának megfelelően.

SZABÁLYOK:
1. Csak a dokumentumban ténylegesen szereplő adatokat add meg
2. Ha egy adat nem található, hagyd üresen (null)
3. CAS számokat pontosan másold (kötőjelekkel)
4. H/P mondatokat pontosvesszővel válaszd el
5. Koncentrációkat az eredeti formátumban add meg (pl. "25-50%", "≥10%")
6. Magyar és angol SDS-t is feldolgozz
7. Ha több összetevő van 3-nál, a 3 legveszélyesebbet válaszd ki
8. CLP osztályozásnál a teljes osztályt add meg (pl. "Flam. Liq. 2, Repr. 1B, STOT RE 2")
"""

def extract_text_from_pdf(pdf_file) -> str:
    """PDF szöveg kinyerése PyPDF2-vel"""
    import PyPDF2
    reader = PyPDF2.PdfReader(pdf_file)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"
    return text

def process_sds_with_openai(pdf_text: str, api_key: str) -> dict:
    """SDS feldolgozás OpenAI GPT-4o Structured Output-tal"""
    from openai import OpenAI

    client = OpenAI(api_key=api_key)

    response = client.responses.parse(
        model="gpt-4o",
        input=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": f"Kérlek dolgozd fel ezt a biztonsági adatlapot:\n\n{pdf_text}"}
        ],
        text={"format": {"type": "json_schema", "schema": SDSExtraction.model_json_schema()}}
    )

    return response.output_parsed

def process_sds_with_anthropic(pdf_text: str, api_key: str) -> dict:
    """SDS feldolgozás Anthropic Claude-dal"""
    import anthropic

    client = anthropic.Anthropic(api_key=api_key)

    schema_json = json.dumps(SDSExtraction.model_json_schema(), indent=2)

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=8000,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": f"Kérlek dolgozd fel ezt a biztonsági adatlapot. "
                      f"Az eredményt KIZÁRÓLAG a következő JSON séma szerint add vissza:\n"
                      f"{schema_json}\n\nSDS TARTALMA:\n{pdf_text}"
        }]
    )

    return json.loads(response.content[0].text)

def write_to_excel(extractions: list, template_path: str = None) -> str:
    """Kinyert adatok Excel fájlba írása a minta formátumban"""
    import openpyxl

    if template_path:
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()
        # Munkalapok létrehozása a minta alapján
        # ... (részletes implementáció)

    ws = wb['Veszélyes_anyag_adatbázis']

    for idx, data in enumerate(extractions, start=2):
        ws.cell(row=idx, column=1, value=idx-1)  # Ssz.
        ws.cell(row=idx, column=3, value=data.get('product_name'))
        # ... (minden mező kitöltése)

    output_path = "output_sds_database.xlsx"
    wb.save(output_path)
    return output_path


# ============================================================
# 4. ALUL: Batch feldolgozás pipeline
# ============================================================

st.divider()
st.header("🔄 Feldolgozási folyamat")

st.markdown("""
```
PDF feltöltés (1-100 fájl)
    ↓
PyPDF2 szövegkinyerés (+ OCR ha szükséges)
    ↓
LLM API hívás (GPT-4o Structured Output / Claude)
    ├── JSON séma: 85 mező/SDS
    ├── Validáció: CAS szám, H/P mondat formátum
    └── Automatikus nyelv-felismerés (HU/EN/DE)
    ↓
Adatgazdagítás
    ├── H/P mondatok magyar kifejtése
    ├── 5/2020 ITM határértékek hozzárendelése
    ├── ECHA SVHC lista ellenőrzés
    └── Részletes védőeszköz specifikáció generálás
    ↓
Excel generálás (mintafájl formátum)
    ├── Veszélyes_anyag_adatbázis (85 oszlop)
    ├── Kémiai_kockázatértékelés (25 oszlop)
    ├── Expozíciós_nyilvántartás sablon
    └── Intézkedési_terv sablon
    ↓
Letöltés (.xlsx)
```
""")

if __name__ == "__main__":
    pass

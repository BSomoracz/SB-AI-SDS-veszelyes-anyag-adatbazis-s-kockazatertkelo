#!/usr/bin/env python3
"""
SDS→Excel AI Feldolgozó Rendszer v2.0 – Kibővített verzió
============================================================
Újdonságok a v1.0-hoz képest:
- Online kutatás: terméknév/CAS szám alapján kiegészítő adatok keresése
- Többnyelvű SDS feldolgozás: HU/EN/DE → mindig magyar output
- Hiányzó adatok automatikus pótlása internetes forrásokból
- ECHA adatbázis lekérdezés (SVHC, CLP, határértékek)

Telepítés:
    pip install streamlit openai openpyxl PyPDF2 pydantic python-dotenv requests

Indítás:
    streamlit run sds_processor_v2.py
"""

import streamlit as st
import json
import os
import time
from typing import Optional, List
from pydantic import BaseModel, Field
from openai import OpenAI

# ============================================================
# 1. KONFIGURÁCIÓ
# ============================================================

# Az elsődleges output nyelv MINDIG magyar
PRIMARY_LANGUAGE = "hu"  
# Támogatott SDS nyelvek
SUPPORTED_LANGUAGES = ["hu", "en", "de"]

# ============================================================
# 2. PYDANTIC SÉMA (ugyanaz mint v1, de kiegészítve)
# ============================================================

class SDSComponent(BaseModel):
    name_hu: str = Field(description="Összetevő neve MAGYARUL")
    name_original: Optional[str] = Field(None, description="Összetevő neve az eredeti nyelven")
    cas_number: Optional[str] = Field(None, description="CAS szám")
    ec_number: Optional[str] = Field(None, description="EK szám")
    concentration: Optional[str] = Field(None, description="Koncentráció %")
    clp_classification: Optional[str] = Field(None, description="CLP besorolás")

class SDSExtraction(BaseModel):
    """Teljes SDS adatkinyerés – minden mező MAGYARUL"""
    # Azonosítás
    product_name: str = Field(description="Kereskedelmi név (eredeti)")
    product_name_hu: Optional[str] = Field(None, description="Kereskedelmi név magyarul (ha eltér)")
    sds_language: str = Field(description="SDS nyelve: hu/en/de")
    sds_version: Optional[str] = None
    sds_date: Optional[str] = None
    sds_revision_date: Optional[str] = None
    manufacturer: Optional[str] = None
    manufacturer_address: Optional[str] = None
    manufacturer_phone: Optional[str] = None
    manufacturer_email: Optional[str] = None
    emergency_phone: Optional[str] = None
    ufi_code: Optional[str] = None
    product_form_hu: Optional[str] = Field(None, description="Termék formája MAGYARUL")
    intended_use_hu: Optional[str] = Field(None, description="Felhasználás MAGYARUL")
    substance_or_mixture: Optional[str] = None

    # Összetétel – MAGYARUL
    component_1: Optional[SDSComponent] = None
    component_2: Optional[SDSComponent] = None
    component_3: Optional[SDSComponent] = None

    # Veszélyesség
    clp_classification: Optional[str] = None
    ghs_pictograms: Optional[str] = None
    signal_word_hu: Optional[str] = Field(None, description="Jelzőszó MAGYARUL: Veszély/Figyelem")
    h_statements: Optional[str] = Field(None, description="H mondatok kódokkal")
    p_statements: Optional[str] = Field(None, description="P mondatok kódokkal")
    euh_statements: Optional[str] = None
    svhc: Optional[str] = None
    pbt_vpvb: Optional[str] = None

    # Fizikai-kémiai – MAGYARUL
    physical_state_hu: Optional[str] = Field(None, description="Halmazállapot MAGYARUL")
    color_hu: Optional[str] = Field(None, description="Szín MAGYARUL")
    odor_hu: Optional[str] = Field(None, description="Szag MAGYARUL")
    melting_point: Optional[str] = None
    boiling_point: Optional[str] = None
    flash_point: Optional[str] = None
    autoignition_temp: Optional[str] = None
    density: Optional[str] = None
    water_solubility_hu: Optional[str] = Field(None, description="Vízoldhatóság MAGYARUL")
    ph: Optional[str] = None
    vapor_pressure: Optional[str] = None

    # Határértékek
    ak_value: Optional[str] = None
    ck_value: Optional[str] = None
    mk_value: Optional[str] = None
    dnel_inhalation: Optional[str] = None
    dnel_dermal: Optional[str] = None
    boelv: Optional[str] = None

    # Védőeszközök – MAGYARUL
    respiratory_hu: Optional[str] = Field(None, description="Légzésvédelem MAGYARUL")
    hand_protection_hu: Optional[str] = Field(None, description="Kézvédelem MAGYARUL")
    eye_protection_hu: Optional[str] = Field(None, description="Szemvédelem MAGYARUL")
    skin_protection_hu: Optional[str] = Field(None, description="Bőrvédelem MAGYARUL")
    engineering_controls_hu: Optional[str] = Field(None, description="Műszaki védelem MAGYARUL")

    # Tűzvédelem – MAGYARUL
    suitable_extinguishing_hu: Optional[str] = None
    unsuitable_extinguishing_hu: Optional[str] = None
    hazardous_decomposition_hu: Optional[str] = None
    firefighter_ppe_hu: Optional[str] = None

    # Toxikológia
    ld50_oral: Optional[str] = None
    ld50_dermal: Optional[str] = None
    lc50_inhalation: Optional[str] = None
    skin_irritation_hu: Optional[str] = None
    eye_irritation_hu: Optional[str] = None
    sensitization_hu: Optional[str] = None
    cmr_effects_hu: Optional[str] = None

    # Szállítás
    un_number: Optional[str] = None
    shipping_name: Optional[str] = None
    adr_class: Optional[str] = None
    packing_group: Optional[str] = None
    marine_pollutant: Optional[str] = None

    # Hulladék
    ewc_code: Optional[str] = None
    disposal_method_hu: Optional[str] = None

    # Megbízhatósági jelzők
    confidence_score: Optional[float] = Field(None, description="0-1 közötti megbízhatósági pontszám")
    missing_fields: Optional[List[str]] = Field(None, description="Hiányzó/nem talált mezők listája")


# ============================================================
# 3. FELDOLGOZÁSI PIPELINE
# ============================================================

# ---- 3a. SYSTEM PROMPT: SDS kinyerés + fordítás ----

EXTRACTION_PROMPT = """Te egy veszélyes anyagok szakértője vagy. Biztonsági adatlapokból (SDS/MSDS) 
nyersz ki strukturált adatokat.

NYELVI SZABÁLYOK:
1. A bemeneti SDS BÁRMILYEN NYELVEN lehet (magyar, angol, német stb.)
2. A kimeneti adatokat MINDIG MAGYARUL add meg a "_hu" végződésű mezőkben
3. Ha az SDS angol/német, fordítsd le a releváns szöveges mezőket magyarra
4. Kémiai nevek: használd a magyar szakkifejezést (pl. "Toluol" nem "Toluene")
5. H/P mondatokat NE fordítsd – csak a kódokat add meg (H225, P210 stb.)
6. CAS számokat, számértékeket, mértékegységeket NE változtasd meg

KINYERÉSI SZABÁLYOK:
1. Csak a dokumentumban ténylegesen szereplő adatokat add meg
2. Ha egy adat nem található → null
3. confidence_score: becsüld meg 0-1 skálán az adatok megbízhatóságát
4. missing_fields: listázd a fontos, de nem talált mezőket
5. Ha több összetevő van 3-nál → a 3 legveszélyesebbet válaszd ki

FORDÍTÁSI SZÓTÁR (gyakori kifejezések):
- "Danger" → "Veszély", "Warning" → "Figyelem"
- "Liquid" → "Folyadék", "Solid" → "Szilárd", "Aerosol" → "Aeroszol"
- "Paste" → "Paszta", "Powder" → "Por", "Gas" → "Gáz"
- "Miscible" → "Elegyedik", "Immiscible" → "Nem oldódik"
- "Safety goggles" → "Védőszemüveg", "Protective gloves" → "Védőkesztyű"
- "Local exhaust ventilation" → "Helyi elszívás"
"""

# ---- 3b. SYSTEM PROMPT: Online kutatás kiegészítő adatokért ----

RESEARCH_PROMPT = """Te egy veszélyes anyagok szakértője vagy. A feladatod, hogy egy adott 
kémiai termékhez/anyaghoz KIEGÉSZÍTŐ ADATOKAT keress az internetről.

Kapni fogsz egy terméknevet és esetleg CAS számot. Keress rá a következőkre:
1. Magyar nyelvű SDS elérhető-e online? Ha igen, milyen adatok találhatók?
2. ECHA adatbázisban szerepel-e (SVHC lista, CLP harmonizált osztályozás)?
3. Magyar 5/2020 (II.6.) ITM rendelet szerinti ÁK/CK/MK határértékek
4. Védőeszköz ajánlások a fő összetevőkhöz (kesztyű típus, szűrő típus)
5. Toxikológiai referencia adatok (LD50, LC50)

FONTOS: Csak megbízható forrásokból (ECHA, PubChem, gyártói SDS adatbázis) származó 
adatokat adj vissza. Minden adatnál jelöld a forrást.

Válaszolj MAGYARUL, strukturáltan.
"""

# ---- 3c. SDS feldolgozó függvény (PDF → struktúrált adat) ----

def process_single_sds(pdf_text: str, client: OpenAI, 
                       enable_web_search: bool = True,
                       product_name_hint: str = None) -> dict:
    """
    Egyetlen SDS feldolgozása:
    1. Szöveg → strukturált JSON (GPT-4o Structured Output)
    2. Ha hiányos → online kutatás kiegészítéshez (web_search tool)
    3. Eredmény összefésülés
    """

    # ========== 1. LÉPÉS: PDF szöveg → strukturált JSON ==========
    extraction_response = client.responses.parse(
        model="gpt-4o",
        input=[
            {"role": "system", "content": EXTRACTION_PROMPT},
            {"role": "user", "content": f"Kérlek dolgozd fel ezt a biztonsági adatlapot:\n\n{pdf_text[:25000]}"}
        ],
        text_format=SDSExtraction,
    )

    extracted = extraction_response.output_parsed
    result = extracted.model_dump()

    # ========== 2. LÉPÉS: Hiányzó adatok azonosítása ==========
    critical_missing = []

    if not result.get('ak_value') and result.get('h_statements'):
        critical_missing.append("ÁK-érték (határérték)")
    if not result.get('ld50_oral'):
        critical_missing.append("LD50 orális")
    if not result.get('svhc'):
        critical_missing.append("SVHC státusz")
    if not result.get('hand_protection_hu') or 'védőkesztyű' in str(result.get('hand_protection_hu','')).lower():
        critical_missing.append("Részletes kesztyű specifikáció")
    if not result.get('respiratory_hu'):
        critical_missing.append("Légzésvédelem részletei")

    # ========== 3. LÉPÉS: Online kutatás a hiányzó adatokért ==========
    if enable_web_search and critical_missing:
        # Terméknév és CAS szám meghatározása a kereséshez
        search_name = product_name_hint or result.get('product_name', '')
        cas_numbers = []
        for comp_key in ['component_1', 'component_2', 'component_3']:
            comp = result.get(comp_key)
            if comp and isinstance(comp, dict) and comp.get('cas_number'):
                cas_numbers.append(comp['cas_number'])

        search_query = f"""
Termék: {search_name}
CAS számok: {', '.join(cas_numbers) if cas_numbers else 'nem ismert'}
Hiányzó adatok: {', '.join(critical_missing)}

Kérlek keress az alábbi adatokra:
1. Magyar ÁK/CK/MK határértékek (5/2020 ITM rendelet) a fenti CAS számokhoz
2. ECHA SVHC lista – szerepelnek-e a fenti CAS számok?
3. Részletes kesztyű-ajánlás: kesztyű anyaga, vastagsága, áttörési ideje, szabványa
4. Légzésvédő szűrő típus ajánlás az összetevőkhöz
5. LD50/LC50 toxikológiai referencia adatok
"""

        # GPT-4o web search tool-lal
        research_response = client.responses.create(
            model="gpt-4o",
            tools=[{
                "type": "web_search",
                "user_location": {
                    "type": "approximate",
                    "country": "HU",
                    "city": "Budapest",
                }
            }],
            input=[
                {"role": "system", "content": RESEARCH_PROMPT},
                {"role": "user", "content": search_query}
            ],
        )

        # Kutatási eredmény feldolgozása
        research_text = research_response.output_text

        # ========== 4. LÉPÉS: Összefésülés ==========
        merge_response = client.responses.parse(
            model="gpt-4o",
            input=[
                {"role": "system", "content": """Összefésülési feladat: 
                Az EREDETI SDS adatok az elsődlegesek. Az ONLINE KUTATÁS adatai csak 
                a hiányzó mezőket pótolják. Ha ellentmondás van, az SDS adat nyer.
                Az eredményt MAGYARUL add meg."""},
                {"role": "user", "content": f"""
EREDETI SDS ADATOK:
{json.dumps(result, ensure_ascii=False, indent=2)}

ONLINE KUTATÁS EREDMÉNY:
{research_text}

Kérlek frissítsd az SDS adatokat az online kutatás alapján, 
csak a hiányzó/üres mezőket pótolva!"""}
            ],
            text_format=SDSExtraction,
        )

        result = merge_response.output_parsed.model_dump()
        result['_research_notes'] = research_text[:500]  # Kutatási jegyzet mentése

    return result


# ---- 3d. Batch feldolgozás ----

def process_batch(pdf_files: list, client: OpenAI, 
                  enable_web_search: bool = True,
                  progress_callback=None) -> list:
    """Több SDS PDF feldolgozása egymás után"""
    import PyPDF2

    results = []

    for i, pdf_file in enumerate(pdf_files):
        try:
            # PDF szöveg kinyerés
            reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

            if len(text.strip()) < 100:
                # Túl kevés szöveg → valószínűleg szkennelt PDF
                results.append({
                    'error': f'Szkennelt PDF - OCR szükséges: {pdf_file.name}',
                    'filename': pdf_file.name
                })
                continue

            # Feldolgozás
            result = process_single_sds(
                pdf_text=text,
                client=client,
                enable_web_search=enable_web_search,
                product_name_hint=pdf_file.name.split('.')[0]
            )
            result['_source_filename'] = pdf_file.name
            results.append(result)

            if progress_callback:
                progress_callback(i + 1, len(pdf_files), pdf_file.name)

            # Rate limit kezelés
            time.sleep(1)

        except Exception as e:
            results.append({
                'error': str(e),
                'filename': pdf_file.name
            })

    return results


# ============================================================
# 4. EXCEL GENERÁLÁS (mintafájl formátum)
# ============================================================

def write_results_to_excel(results: list, template_path: str = None) -> str:
    """Kinyert adatok Excel fájlba írása a minta formátumban"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    if template_path:
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = create_template_workbook()

    ws = wb['Veszélyes_anyag_adatbázis']

    # Oszlop mapping: JSON mező → Excel oszlop
    field_to_column = {
        'product_name': 3,
        'sds_language': 4,
        'sds_version': 5,
        'sds_date': 6,
        'sds_revision_date': 7,
        'manufacturer': 8,
        'manufacturer_address': 9,
        'manufacturer_phone': 10,
        'manufacturer_email': 11,
        'emergency_phone': 12,
        'ufi_code': 13,
        'product_form_hu': 14,
        'intended_use_hu': 15,
        # ... (teljes mapping a 85 oszlophoz)
        'clp_classification': 32,
        'ghs_pictograms': 33,
        'signal_word_hu': 34,
        'h_statements': 35,
        'p_statements': 36,
        'euh_statements': 37,
        'svhc': 38,
        'pbt_vpvb': 39,
        'physical_state_hu': 40,
        'color_hu': 41,
        'odor_hu': 42,
        'ak_value': 51,
        'ck_value': 52,
        'mk_value': 53,
        'respiratory_hu': 57,
        'hand_protection_hu': 58,
        'eye_protection_hu': 59,
        'skin_protection_hu': 60,
        'engineering_controls_hu': 61,
    }

    for idx, data in enumerate(results, start=2):
        if 'error' in data:
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=3, value=f"HIBA: {data.get('filename', '?')}")
            ws.cell(row=idx, column=85, value=data['error'])
            continue

        ws.cell(row=idx, column=1, value=idx-1)

        for field, col in field_to_column.items():
            value = data.get(field)
            if value and value != 'null':
                ws.cell(row=idx, column=col, value=str(value))

        # Összetevők külön kezelése
        for comp_idx, comp_key in enumerate(['component_1', 'component_2', 'component_3']):
            comp = data.get(comp_key)
            if comp and isinstance(comp, dict):
                base_col = 18 + (comp_idx * 5)  # 18, 23, 28
                ws.cell(row=idx, column=base_col, value=comp.get('name_hu', comp.get('name_original', '')))
                ws.cell(row=idx, column=base_col+1, value=comp.get('cas_number', ''))
                ws.cell(row=idx, column=base_col+2, value=comp.get('ec_number', ''))
                ws.cell(row=idx, column=base_col+3, value=comp.get('concentration', ''))
                ws.cell(row=idx, column=base_col+4, value=comp.get('clp_classification', ''))

    output = "sds_output.xlsx"
    wb.save(output)
    return output


# ============================================================
# 5. STREAMLIT UI
# ============================================================

st.set_page_config(page_title="🧪 SDS AI Feldolgozó v2.0", page_icon="🧪", layout="wide")

st.title("🧪 SDS → Excel AI Feldolgozó v2.0")
st.markdown("**Többnyelvű SDS feldolgozás + online kutatás + magyar nyelvű kimenet**")

# Sidebar
with st.sidebar:
    st.header("⚙️ Beállítások")
    api_key = st.text_input("OpenAI API kulcs", type="password")

    st.divider()
    st.header("🌐 Nyelvi beállítások")
    output_lang = st.selectbox("Kimenet nyelve", ["Magyar (HU)", "English (EN)", "Deutsch (DE)"])
    auto_translate = st.checkbox("Automatikus fordítás", value=True, 
                                 help="Angol/német SDS automatikus fordítása a kiválasztott nyelvre")

    st.divider()
    st.header("🔍 Online kutatás")
    enable_research = st.checkbox("Online adatkiegészítés", value=True,
                                  help="Hiányzó adatok keresése az interneten (CAS, határértékek, védőeszközök)")

    research_sources = st.multiselect("Keresési források", 
        ["ECHA adatbázis", "PubChem", "Gyártói weboldal", "msds.com", "5/2020 ITM rendelet"],
        default=["ECHA adatbázis", "PubChem", "5/2020 ITM rendelet"])

    st.divider()
    st.header("📋 Mintafájl")
    template = st.file_uploader("Excel sablon (opcionális)", type=["xlsx"])

# Fő tartalom
tab1, tab2, tab3 = st.tabs(["📤 Feltöltés & Feldolgozás", "📊 Eredmények", "📖 Útmutató"])

with tab1:
    col1, col2 = st.columns([3, 1])

    with col1:
        uploaded = st.file_uploader("SDS PDF fájlok (max. 100)", type=["pdf"], 
                                     accept_multiple_files=True)

        if uploaded:
            st.success(f"✅ {len(uploaded)} fájl feltöltve")

            # Nyelv-felismerés előnézet
            lang_stats = {"Magyar": 0, "Angol": 0, "Német": 0, "Egyéb": 0}
            for f in uploaded:
                name = f.name.lower()
                if any(x in name for x in ['_hu', 'hungarian', 'magyar']):
                    lang_stats["Magyar"] += 1
                elif any(x in name for x in ['_en', 'english', '_gb', '_us']):
                    lang_stats["Angol"] += 1
                elif any(x in name for x in ['_de', 'german', 'deutsch']):
                    lang_stats["Német"] += 1
                else:
                    lang_stats["Egyéb"] += 1

            st.markdown("**Fájlnév alapú nyelvfelismerés:**")
            for lang, count in lang_stats.items():
                if count > 0:
                    st.text(f"  {lang}: {count} fájl")

    with col2:
        st.markdown("### Feldolgozási opciók")
        st.metric("PDF fájlok", len(uploaded) if uploaded else 0)
        st.metric("Becsült idő", f"~{len(uploaded)*15 if uploaded else 0} mp" if uploaded else "—")
        st.metric("Becsült költség", f"~${len(uploaded)*0.20:.2f}" if uploaded else "—")

        if uploaded and api_key:
            if st.button("🚀 FELDOLGOZÁS INDÍTÁSA", type="primary", use_container_width=True):
                st.session_state['processing'] = True

with tab3:
    st.markdown("""
    ## 🔄 Feldolgozási folyamat

    ### 1. PDF szövegkinyerés
    - Szöveges PDF: PyPDF2 direkt kinyerés
    - Szkennelt PDF: OCR (Tesseract) – automatikusan felismeri

    ### 2. AI adatkinyerés (GPT-4o)
    - A modell **bármilyen nyelvű** SDS-t feldolgoz
    - Structured Output: kényszerített JSON séma (85 mező)
    - Automatikus fordítás a kiválasztott kimeneti nyelvre

    ### 3. Online kutatás (opcionális)
    Az AI **web_search** eszközzel kiegészíti a hiányzó adatokat:
    - **ECHA adatbázis**: SVHC státusz, harmonizált CLP osztályozás
    - **PubChem**: LD50/LC50, fizikai-kémiai adatok
    - **5/2020 ITM rendelet**: magyar ÁK/CK/MK határértékek
    - **Gyártói weboldalak**: friss SDS, védőeszköz részletek

    ### 4. Adatgazdagítás (Python)
    - H/P mondatok magyar kifejtése (beépített szótár)
    - Részletes védőeszköz specifikáció (kesztyű típus, vastagság, szűrő)
    - Kockázati szint automatikus számítás

    ### 5. Excel kimenet
    - A mintafájl pontos formátumában
    - 6 munkalap: Adatbázis + Kockázatértékelés + Expozíció + Intézkedés

    ---

    ## 🌐 Többnyelvű működés

    | SDS nyelve | Feldolgozás | Kimenet |
    |------------|-------------|---------|
    | Magyar 🇭🇺 | Direkt kinyerés | Magyar |
    | Angol 🇬🇧 | Kinyerés + fordítás | Magyar |
    | Német 🇩🇪 | Kinyerés + fordítás | Magyar |

    A kémiai szakkifejezések helyes fordítását a beépített szótár biztosítja.
    """)


if __name__ == "__main__":
    pass

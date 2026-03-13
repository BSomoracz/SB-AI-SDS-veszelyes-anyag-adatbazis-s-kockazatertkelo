#!/usr/bin/env python3
"""
SDS → Excel AI Feldolgozó Rendszer v3.2 – FIZETŐS VERZIÓ (Stripe + Barion)
================================================================================
24 EU nyelv | 6 munkalap | H/P kifejtés | Védőeszköz spec. | Kockázatértékelés
Stripe (st-paywall) + Barion Smart Gateway párhuzamos fizetési rendszer
"""

import streamlit as st
import json, time, os, re
import os, subprocess, sys
from typing import Optional, List, Dict
from datetime import datetime
from io import BytesIO

st.set_page_config(page_title="🧪 SDS AI Feldolgozó v3.2", page_icon="🧪", layout="wide")

for key in ['results', 'risk_results', 'processing_done']:
    if key not in st.session_state:
        st.session_state[key] = [] if key != 'processing_done' else False

# ============================================================
# FIZETÉSI RENDSZER – Stripe + Barion párhuzamosan
# ============================================================

def check_barion_payment():
    """Barion fizetés státusz ellenőrzése a Barion API-n keresztül."""
    import requests
    
    barion_poskey = st.secrets.get("barion_poskey", "")
    barion_api_url = st.secrets.get("barion_api_url", "https://api.test.barion.com")
    
    # Ha van payment_id a query paraméterekben (Barion redirect után)
    query_params = st.query_params
    barion_payment_id = query_params.get("barionPaymentId", "")
    
    if barion_payment_id and barion_poskey:
        try:
            resp = requests.get(
                f"{barion_api_url}/v4/payment/{barion_payment_id}/paymentstate",
                headers={"x-pos-key": barion_poskey},
                timeout=10
            )
            if resp.status_code == 200:
                data = resp.json()
                if data.get("Status") == "Succeeded":
                    st.session_state["barion_paid"] = True
                    st.session_state["barion_payment_id"] = barion_payment_id
                    st.session_state["barion_payer_email"] = data.get("Transactions", [{}])[0].get("PayerEmail", "")
                    return True
        except Exception:
            pass
    
    return st.session_state.get("barion_paid", False)


def start_barion_payment():
    """Barion fizetés indítása – átirányítás a Barion Smart Gateway-re."""
    import requests
    
    barion_poskey = st.secrets.get("barion_poskey", "")
    barion_api_url = st.secrets.get("barion_api_url", "https://api.test.barion.com")
    barion_payee = st.secrets.get("barion_payee_email", "")
    barion_redirect_url = st.secrets.get("barion_redirect_url", "")
    barion_callback_url = st.secrets.get("barion_callback_url", "")
    
    # Egyedi fizetési azonosító
    payment_request_id = f"SDS-{datetime.now().strftime('%Y%m%d%H%M%S')}-{os.urandom(4).hex()}"
    
    # Ár a secrets-ből
    price = float(st.secrets.get("service_price_huf", "2990"))
    
    payload = {
        "POSKey": barion_poskey,
        "PaymentType": "Immediate",
        "PaymentRequestId": payment_request_id,
        "GuestCheckout": True,
        "FundingSources": ["All"],
        "Currency": "HUF",
        "Locale": "hu-HU",
        "RedirectUrl": barion_redirect_url,
        "CallbackUrl": barion_callback_url,
        "Transactions": [
            {
                "POSTransactionId": f"T-{payment_request_id}",
                "Payee": barion_payee,
                "Total": price,
                "Comment": "SDS AI Feldolgozó – egyszeri használat",
                "Items": [
                    {
                        "Name": "SDS AI Feldolgozó v3.2 – egyszeri használat",
                        "Description": "Biztonsági adatlap (SDS/MSDS) feldolgozás AI-val, Excel kimenet",
                        "Quantity": 1,
                        "Unit": "db",
                        "UnitPrice": price,
                        "ItemTotal": price,
                    }
                ]
            }
        ]
    }
    
    try:
        resp = requests.post(
            f"{barion_api_url}/v2/Payment/Start",
            json=payload,
            headers={"Content-Type": "application/json"},
            timeout=15
        )
        if resp.status_code == 200:
            data = resp.json()
            gateway_url = data.get("GatewayUrl", "")
            if gateway_url:
                return gateway_url
            else:
                errors = data.get("Errors", [])
                error_msg = "; ".join([e.get("Description", "") for e in errors]) if errors else "Ismeretlen hiba"
                st.error(f"❌ Barion hiba: {error_msg}")
                return None
        else:
            st.error(f"❌ Barion API hiba: HTTP {resp.status_code}")
            return None
    except Exception as e:
        st.error(f"❌ Barion kapcsolódási hiba: {e}")
        return None


def show_payment_wall():
    """
    Fizetési fal megjelenítése – a felhasználó választhat Stripe vagy Barion között.
    Visszatérési érték: True ha fizetett, False ha nem.
    """
    # 1) Barion ellenőrzés (redirect után)
    if check_barion_payment():
        return True
    
    # 2) Stripe ellenőrzés (st-paywall)
    stripe_enabled = bool(st.secrets.get("stripe_api_key_test", "") or st.secrets.get("stripe_api_key", ""))
    barion_enabled = bool(st.secrets.get("barion_poskey", ""))
    
    if stripe_enabled:
        try:
            from st_paywall import add_auth
            # st-paywall – ha a felhasználó be van jelentkezve ÉS előfizető
            if hasattr(st, 'user') and hasattr(st.user, 'is_logged_in') and st.user.is_logged_in:
                if st.session_state.get("user_subscribed", False):
                    return True
        except ImportError:
            stripe_enabled = False
    
    # ============================================================
    # FIZETÉSI FELÜLET
    # ============================================================
    
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; padding: 32px 0 16px 0;">
        <h2 style="margin-bottom: 8px;">🔒 Fizetős szolgáltatás</h2>
        <p style="font-size: 1.1rem; opacity: 0.8;">
            Az SDS AI Feldolgozó használatához fizetés szükséges.<br>
            Válassza ki a kívánt fizetési módot.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    price_huf = st.secrets.get("service_price_huf", "2990")
    
    st.markdown(f"""
    <div style="text-align: center; padding: 16px; background: linear-gradient(135deg, #0A2540 0%, #1a3a5c 100%); 
                border-radius: 12px; margin: 16px auto; max-width: 500px; color: white;">
        <div style="font-size: 2.5rem; font-weight: 800; margin-bottom: 4px;">{price_huf} Ft</div>
        <div style="font-size: 0.95rem; opacity: 0.8;">egyszeri használat / alkalom</div>
        <div style="margin-top: 12px; font-size: 0.85rem; opacity: 0.7;">
            ✓ Korlátlan PDF feldolgozás &nbsp; ✓ 24 EU nyelv &nbsp; ✓ Excel kimenet &nbsp; ✓ Kockázatértékelés
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    # --- STRIPE oszlop ---
    with col1:
        st.markdown("""
        <div style="text-align: center; padding: 20px; border: 2px solid #635BFF; border-radius: 12px; 
                    background: rgba(99,91,255,0.05); height: 100%;">
            <div style="font-size: 1.5rem; font-weight: 700; color: #635BFF; margin-bottom: 8px;">
                💳 Stripe
            </div>
            <div style="font-size: 0.85rem; opacity: 0.7; margin-bottom: 12px;">
                Nemzetközi bankkártyás fizetés<br>
                Visa, Mastercard, Apple Pay, Google Pay
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        if stripe_enabled:
            # Streamlit bejelentkezés + st-paywall
            if not (hasattr(st, 'user') and hasattr(st.user, 'is_logged_in') and st.user.is_logged_in):
                if st.button("🔑 Bejelentkezés (Google)", key="stripe_login", use_container_width=True):
                    st.login()
            else:
                try:
                    from st_paywall import add_auth
                    add_auth(
                        required=False,
                        subscription_button_text="💳 Stripe – Fizetés",
                        button_color="#635BFF",
                        use_sidebar=False
                    )
                    if st.session_state.get("user_subscribed", False):
                        return True
                except ImportError:
                    stripe_link = st.secrets.get("stripe_link", "") or st.secrets.get("stripe_link_test", "")
                    if stripe_link:
                        st.link_button("💳 Stripe – Fizetés", stripe_link, use_container_width=True)
        else:
            st.info("ℹ️ Stripe jelenleg nincs konfigurálva.")
    
    # --- BARION oszlop ---
    with col2:
        st.markdown("""
        <div style="text-align: center; padding: 20px; border: 2px solid #3FB553; border-radius: 12px;
                    background: rgba(63,181,83,0.05); height: 100%;">
            <div style="font-size: 1.5rem; font-weight: 700; color: #3FB553; margin-bottom: 8px;">
                🏦 Barion
            </div>
            <div style="font-size: 0.85rem; opacity: 0.7; margin-bottom: 12px;">
                Magyar fizetési rendszer (HUF)<br>
                Bankkártya, Barion tárca, átutalás
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        if barion_enabled:
            if st.button("🏦 Barion – Fizetés", key="barion_pay", use_container_width=True, type="primary"):
                with st.spinner("Barion fizetés indítása..."):
                    gateway_url = start_barion_payment()
                    if gateway_url:
                        st.markdown(f"""
                        <meta http-equiv="refresh" content="0;url={gateway_url}">
                        <p>Átirányítás a Barion fizetési oldalra... 
                        <a href="{gateway_url}" target="_self">Kattintson ide, ha nem történik automatikus átirányítás.</a></p>
                        """, unsafe_allow_html=True)
                        st.stop()
        else:
            st.info("ℹ️ Barion jelenleg nincs konfigurálva.")
    
    st.markdown("""
    <div style="text-align: center; padding: 16px 0; font-size: 0.8rem; opacity: 0.5;">
        🔒 Minden fizetés biztonságos, titkosított csatornán történik.<br>
        A szolgáltatás igénybevételéhez válassza ki a kívánt fizetési módot.
    </div>
    """, unsafe_allow_html=True)
    
    st.stop()
    return False


# ============================================================
# NYELVI SZÓTÁRAK
# ============================================================
LANGUAGES = {
    "🇭🇺 Magyar": "hu", "🇬🇧 English": "en", "🇩🇪 Deutsch": "de",
    "🇫🇷 Français": "fr", "🇪🇸 Español": "es", "🇮🇹 Italiano": "it",
    "🇵🇱 Polski": "pl", "🇷🇴 Română": "ro", "🇳🇱 Nederlands": "nl",
    "🇵🇹 Português": "pt", "🇨🇿 Čeština": "cs", "🇸🇰 Slovenčina": "sk",
    "🇭🇷 Hrvatski": "hr", "🇧🇬 Български": "bg", "🇸🇪 Svenska": "sv",
    "🇩🇰 Dansk": "da", "🇫🇮 Suomi": "fi", "🇬🇷 Ελληνικά": "el",
    "🇸🇮 Slovenščina": "sl", "🇪🇪 Eesti": "et", "🇱🇻 Latviešu": "lv",
    "🇱🇹 Lietuvių": "lt", "🇮🇪 Gaeilge": "ga", "🇲🇹 Malti": "mt",
}

LANG_NAMES = {
    "hu": "magyar", "en": "English", "de": "Deutsch", "fr": "français",
    "es": "español", "it": "italiano", "pl": "polski", "ro": "română",
    "nl": "Nederlands", "pt": "português", "cs": "čeština", "sk": "slovenčina",
    "hr": "hrvatski", "bg": "български", "sv": "svenska", "da": "dansk",
    "fi": "suomi", "el": "ελληνικά", "sl": "slovenščina", "et": "eesti",
    "lv": "latviešu", "lt": "lietuvių", "ga": "Gaeilge", "mt": "Malti",
}

LANG_STRINGS = {
    "hu": {
        "main_title": "VESZÉLYES ANYAGOK NYILVÁNTARTÁSA ÉS KÉMIAI KOCKÁZATÉRTÉKELÉS",
        "prepared_by": "Készítette: AI asszisztens a feltöltött biztonsági adatlapok (SDS/MSDS) alapján",
        "prep_date": "Készítés dátuma", "processed_count": "Feldolgozott biztonsági adatlapok száma",
        "legal_bg": "JOGSZABÁLYI HÁTTÉR:", "sheets_content": "MUNKALAPOK TARTALMA:",
        "sheet_names": ["Útmutató", "Segédtáblák", "Veszélyes_anyag_adatbázis", "Kémiai_kockázatértékelés", "Expozíciós_nyilvántartás", "Intézkedési_terv"],
        "sheet_desc": ["Útmutató - Ez a munkalap", "Veszélyes_anyag_adatbázis - Teljes nyilvántartás az összes SDS adattal", "Kémiai_kockázatértékelés - Kockázatértékelés 4x4 mátrix alapján", "Expozíciós_nyilvántartás - Mvt. 63/A. § szerinti munkavállalói nyilvántartás", "Intézkedési_terv - Kockázatcsökkentő intézkedések nyomon követése", "Segédtáblák - Kockázati mátrix, GHS piktogramok, skálák"],
        "markings": 'JELÖLÉSEK: Az "X" karakter piros háttérrel jelöli azokat a mezőket, amelyek szükségesek lennének, de nem találhatók az adott SDS-ben.',
        "empty_cells": "Az üres cellák azt jelentik, hogy az adat nem releváns az adott terméknél.",
        "risk_matrix_title": "KOCKÁZATI MÁTRIX (Valószínűség × Súlyosság)",
        "severity": ["Súlyosság 1\n(Elhanyagolható)", "Súlyosság 2\n(Csekély)", "Súlyosság 3\n(Közepes)", "Súlyosság 4\n(Súlyos)"],
        "probability": ["Valószínűség 4 (Nagyon valószínű)", "Valószínűség 3 (Valószínű)", "Valószínűség 2 (Lehetséges)", "Valószínűség 1 (Nem valószínű)"],
        "risk_levels_title": "KOCKÁZATI SZINTEK:",
        "risk_levels": ["1-2: ELFOGADHATÓ (zöld)", "3-4: TOLERÁLHATÓ (sárga) - intézkedés szükséges", "5-9: JELENTŐS (narancs) - sürgős intézkedés", "10-16: ELFOGADHATATLAN (piros) - azonnali intézkedés / tevékenység leállítása"],
        "ghs_title": "GHS PIKTOGRAMOK:",
        "ghs_symbols": ["Robbanó bomba", "Láng", "Láng kör felett", "Gázpalack", "Maró hatás", "Koponya", "Felkiáltójel", "Egészségi veszély", "Környezet"],
        "ghs_desc": ["Robbanóanyagok", "Tűzveszélyes anyagok", "Oxidáló anyagok", "Nyomás alatt lévő gázok", "Fémekre korrozív; bőrmarás; szemkárosodás", "Akut toxicitás (halálos/mérgező)", "Irritáció; szenzibilizáció; akut tox. 4; narkózis", "CMR; STOT; aspiráció; szenzibilizáció (légúti)", "Vízi környezetre veszélyes"],
        "prob_scale_title": "VALÓSZÍNŰSÉGI SKÁLA:",
        "prob_scale": [("1 - Nem valószínű", "Ritka expozíció, hatékony védelem, zárt rendszer"), ("2 - Lehetséges", "Alkalmi expozíció, részleges védelem"), ("3 - Valószínű", "Rendszeres expozíció, hiányos védelem"), ("4 - Nagyon valószínű", "Folyamatos expozíció, védelem nélkül")],
        "sev_scale_title": "SÚLYOSSÁGI SKÁLA:",
        "sev_scale": [("1 - Elhanyagolható", "Enyhe, reverzibilis hatás (irritáció)"), ("2 - Csekély", "Reverzibilis egészségkárosodás"), ("3 - Közepes", "Súlyos, potenciálisan irreverzibilis hatás"), ("4 - Súlyos", "Halálos/maradandó károsodás, CMR hatás")],
        "db_headers": ["Ssz.", "Termék kategória", "Kereskedelmi név", "SDS nyelve", "SDS verziószám", "SDS kiadás dátuma", "SDS felülvizsgálat dátuma", "Gyártó/Szállító", "Gyártó címe", "Gyártó tel.", "Gyártó e-mail", "Sürgősségi tel.", "UFI kód", "Termék forma", "Felhasználás", "Felhasználási kategória", "Anyag/Keverék", "Fő összetevő 1 - név", "Fő összetevő 1 - CAS", "Fő összetevő 1 - EK szám", "Fő összetevő 1 - konc. %", "Fő összetevő 1 - CLP osztály", "Fő összetevő 2 - név", "Fő összetevő 2 - CAS", "Fő összetevő 2 - EK szám", "Fő összetevő 2 - konc. %", "Fő összetevő 2 - CLP osztály", "Fő összetevő 3 - név", "Fő összetevő 3 - CAS", "Fő összetevő 3 - konc. %", "Fő összetevő 3 - CLP osztály", "CLP osztályozás (keverék)", "GHS piktogram kódok", "Jelzőszó", "H mondatok", "P mondatok", "EUH mondatok", "SVHC anyag", "PBT/vPvB", "Halmazállapot", "Szín", "Szag", "Olvadáspont (°C)", "Forráspont (°C)", "Lobbanáspont (°C)", "Gyulladási hőm. (°C)", "Sűrűség (g/cm³)", "Vízoldhatóság", "pH", "Gőznyomás", "ÁK-érték (mg/m³)", "CK-érték (mg/m³)", "MK-érték (mg/m³)", "DNEL munkás inhaláció", "DNEL munkás dermális", "BOELV (EU) mg/m³", "Légzésvédelem", "Kézvédelem", "Szemvédelem", "Bőrvédelem", "Műszaki védelem", "Megfelelő oltóanyag", "Nem megfelelő oltóanyag", "Veszélyes bomlástermékek", "Tűzoltói védőfelszerelés", "Akut tox. orális LD50", "Akut tox. dermális LD50", "Akut tox. inhal. LC50", "Bőrirritáció", "Szemirritáció", "Szenzibilizáció", "CMR hatások", "UN szám", "Szállítási megnevezés", "ADR osztály", "Csomagolási csop.", "Tengeri szenny.", "EWC kód", "Hulladékkezelés", "Felhasználás helye", "Felhasznált mennyiség/év", "Felhasználás gyakorisága", "Expozíció módja", "Érintett munkavállalók száma", "Megjegyzés"],
        "risk_headers": ["Ssz.", "Kereskedelmi név", "Fő veszélyes összetevő", "CLP osztályozás", "H mondatok", "P mondatok", "Expozíció módja", "Expozíció gyakorisága", "Expozíció időtartam", "Érintett testrész", "Védelem megléte", "Egyéni védőeszköz specifikáció", "Valószínűség (1-4)", "Súlyosság (1-4)", "Kockázat (VxS)", "Kockázati szint", "Szükséges intézkedés", "BEM vizsgálat szükséges", "Munkáltatói expozíciós nyilvántartás vezetése kötelező", "Intézkedés határideje", "Felelős", "Intézkedés utáni valószínűség", "Intézkedés utáni súlyosság", "Maradék kockázat", "Maradék kockázati szint", "Értékelő neve", "Értékelés dátuma", "Felülvizsgálat dátuma", "Megjegyzés"],
        "exp_headers": ["Ssz.", "Munkavállaló neve", "Születési hely és idő", "Anyja neve", "Munkakör", "Munkahely/telephely", "Veszélyes anyag kereskedelmi neve", "Veszélyes anyag CAS száma", "Expozíció módja", "Napi expozíciós idő (óra)", "Heti expozíciós idő (óra)", "Éves expozíciós idő (óra)", "Mért expozíciós koncentráció (mg/m³)", "ÁK/CK határérték (mg/m³)", "Alkalmazott védőeszköz", "Munkaegészségügyi vizsgálat", "Nyilvántartás kezdete", "Megjegyzés"],
        "exp_note": "Mvt. 63/A. § szerinti nyilvántartás - A munkáltató tölti ki munkavállalónként!",
        "action_headers": ["Ssz.", "Veszélyes anyag", "Kockázati szint", "Szükséges intézkedés", "Felelős", "Határidő", "Státusz", "Befejezés dátuma", "Megjegyzés"],
        "use_location": "Termelés", "company_fills": "Vállalat tölti ki!", "employer": "Munkáltató", "in_progress": "Folyamatban",
    },
    "en": {
        "main_title": "HAZARDOUS SUBSTANCES REGISTRY AND CHEMICAL RISK ASSESSMENT",
        "prepared_by": "Prepared by: AI assistant based on uploaded Safety Data Sheets (SDS/MSDS)",
        "prep_date": "Preparation date", "processed_count": "Number of processed safety data sheets",
        "legal_bg": "LEGAL BACKGROUND:", "sheets_content": "WORKSHEET CONTENTS:",
        "sheet_names": ["Guide", "Reference_Tables", "Hazardous_Substance_DB", "Chemical_Risk_Assessment", "Exposure_Registry", "Action_Plan"],
        "sheet_desc": ["Guide - This worksheet", "Hazardous_Substance_DB - Complete registry with all SDS data", "Chemical_Risk_Assessment - Risk assessment based on 4x4 matrix", "Exposure_Registry - Employee exposure registry per legislation", "Action_Plan - Risk reduction measures tracking", "Reference_Tables - Risk matrix, GHS pictograms, scales"],
        "markings": 'MARKINGS: "X" with red background indicates required fields not found in the SDS.',
        "empty_cells": "Empty cells mean the data is not relevant for the given product.",
        "risk_matrix_title": "RISK MATRIX (Probability × Severity)",
        "severity": ["Severity 1\n(Negligible)", "Severity 2\n(Minor)", "Severity 3\n(Moderate)", "Severity 4\n(Severe)"],
        "probability": ["Probability 4 (Very likely)", "Probability 3 (Likely)", "Probability 2 (Possible)", "Probability 1 (Unlikely)"],
        "risk_levels_title": "RISK LEVELS:",
        "risk_levels": ["1-2: ACCEPTABLE (green)", "3-4: TOLERABLE (yellow) - action required", "5-9: SIGNIFICANT (orange) - urgent action", "10-16: UNACCEPTABLE (red) - immediate action / stop activity"],
        "ghs_title": "GHS PICTOGRAMS:",
        "ghs_symbols": ["Exploding bomb", "Flame", "Flame over circle", "Gas cylinder", "Corrosion", "Skull & crossbones", "Exclamation mark", "Health hazard", "Environment"],
        "ghs_desc": ["Explosives", "Flammable", "Oxidizers", "Gases under pressure", "Corrosive to metals; skin corrosion; eye damage", "Acute toxicity (fatal/toxic)", "Irritation; sensitization; acute tox. 4; narcosis", "CMR; STOT; aspiration; respiratory sensitization", "Aquatic hazard"],
        "prob_scale_title": "PROBABILITY SCALE:",
        "prob_scale": [("1 - Unlikely", "Rare exposure, effective protection, closed system"), ("2 - Possible", "Occasional exposure, partial protection"), ("3 - Likely", "Regular exposure, insufficient protection"), ("4 - Very likely", "Continuous exposure, no protection")],
        "sev_scale_title": "SEVERITY SCALE:",
        "sev_scale": [("1 - Negligible", "Mild, reversible (irritation)"), ("2 - Minor", "Reversible health damage"), ("3 - Moderate", "Severe, potentially irreversible"), ("4 - Severe", "Fatal/permanent, CMR effect")],
        "db_headers": ["No.", "Product category", "Trade name", "SDS language", "SDS version", "SDS issue date", "SDS revision date", "Manufacturer/Supplier", "Address", "Phone", "Email", "Emergency phone", "UFI code", "Product form", "Intended use", "Use category", "Substance/Mixture", "Component 1 - name", "Component 1 - CAS", "Component 1 - EC", "Component 1 - conc.%", "Component 1 - CLP", "Component 2 - name", "Component 2 - CAS", "Component 2 - EC", "Component 2 - conc.%", "Component 2 - CLP", "Component 3 - name", "Component 3 - CAS", "Component 3 - conc.%", "Component 3 - CLP", "CLP classification (mixture)", "GHS pictograms", "Signal word", "H statements", "P statements", "EUH statements", "SVHC", "PBT/vPvB", "Physical state", "Colour", "Odour", "Melting pt (°C)", "Boiling pt (°C)", "Flash pt (°C)", "Auto-ign. (°C)", "Density (g/cm³)", "Water solubility", "pH", "Vapour pressure", "OEL-TWA (mg/m³)", "OEL-STEL (mg/m³)", "OEL-C (mg/m³)", "DNEL inhalation", "DNEL dermal", "BOELV (EU) mg/m³", "Respiratory PPE", "Hand protection", "Eye protection", "Skin protection", "Engineering controls", "Suitable extinguishing", "Unsuitable extinguishing", "Hazardous decomposition", "Firefighter PPE", "Oral LD50", "Dermal LD50", "Inhal. LC50", "Skin irritation", "Eye irritation", "Sensitization", "CMR effects", "UN number", "Shipping name", "ADR class", "Packing group", "Marine pollutant", "EWC code", "Waste disposal", "Place of use", "Annual quantity", "Frequency", "Exposure route", "Workers exposed", "Notes"],
        "risk_headers": ["No.", "Trade name", "Main hazardous component", "CLP classification", "H statements", "P statements", "Exposure route", "Frequency", "Duration", "Affected body parts", "Protection present", "PPE specification", "Probability (1-4)", "Severity (1-4)", "Risk (PxS)", "Risk level", "Required action", "BEM required", "Exposure registry required", "Deadline", "Responsible", "Post-action probability", "Post-action severity", "Residual risk", "Residual risk level", "Assessor", "Assessment date", "Review date", "Notes"],
        "exp_headers": ["No.", "Employee name", "Place/date of birth", "Mother's name", "Job title", "Workplace", "Substance trade name", "CAS no.", "Exposure route", "Daily exposure (h)", "Weekly exposure (h)", "Annual exposure (h)", "Measured conc. (mg/m³)", "OEL (mg/m³)", "PPE applied", "Health examination", "Registry start", "Notes"],
        "exp_note": "Exposure registry per legislation - To be completed by employer per employee!",
        "action_headers": ["No.", "Substance", "Risk level", "Required action", "Responsible", "Deadline", "Status", "Completion date", "Notes"],
        "use_location": "Production", "company_fills": "Company to fill!", "employer": "Employer", "in_progress": "In progress",
    },
    "de": {
        "main_title": "GEFAHRSTOFFVERZEICHNIS UND CHEMISCHE GEFÄHRDUNGSBEURTEILUNG",
        "prepared_by": "Erstellt von: KI-Assistent auf Basis hochgeladener Sicherheitsdatenblätter (SDB)",
        "prep_date": "Erstellungsdatum", "processed_count": "Anzahl verarbeiteter Sicherheitsdatenblätter",
        "legal_bg": "RECHTSGRUNDLAGE:", "sheets_content": "INHALT DER ARBEITSBLÄTTER:",
        "sheet_names": ["Anleitung", "Hilfstabellen", "Gefahrstoff_Datenbank", "Gefährdungsbeurteilung", "Expositionsverzeichnis", "Maßnahmenplan"],
        "sheet_desc": ["Anleitung - Dieses Arbeitsblatt", "Gefahrstoff_Datenbank - Vollständiges Verzeichnis", "Gefährdungsbeurteilung - 4x4-Matrix", "Expositionsverzeichnis - Mitarbeiter-Exposition", "Maßnahmenplan - Risikominderung", "Hilfstabellen - Matrix, GHS, Skalen"],
        "markings": 'KENNZEICHNUNGEN: "X" mit rotem Hintergrund = Feld sollte vorhanden sein, fehlt aber im SDB.',
        "empty_cells": "Leere Zellen = Daten nicht relevant für dieses Produkt.",
        "risk_matrix_title": "RISIKOMATRIX (Wahrscheinlichkeit × Schweregrad)",
        "severity": ["Schweregrad 1\n(Vernachlässigbar)", "Schweregrad 2\n(Gering)", "Schweregrad 3\n(Mittel)", "Schweregrad 4\n(Schwer)"],
        "probability": ["Wahrsch. 4 (Sehr wahrscheinlich)", "Wahrsch. 3 (Wahrscheinlich)", "Wahrsch. 2 (Möglich)", "Wahrsch. 1 (Unwahrscheinlich)"],
        "risk_levels_title": "RISIKOSTUFEN:",
        "risk_levels": ["1-2: AKZEPTABEL (grün)", "3-4: TOLERIERBAR (gelb) - Maßnahmen nötig", "5-9: ERHEBLICH (orange) - dringend", "10-16: INAKZEPTABEL (rot) - sofort / Stopp"],
        "ghs_title": "GHS-PIKTOGRAMME:",
        "ghs_symbols": ["Explodierende Bombe", "Flamme", "Flamme über Kreis", "Gasflasche", "Ätzwirkung", "Totenkopf", "Ausrufezeichen", "Gesundheitsgefahr", "Umwelt"],
        "ghs_desc": ["Explosive Stoffe", "Entzündbar", "Oxidierend", "Gase unter Druck", "Korrosiv; Hautverätzung; Augenschädigung", "Akute Toxizität", "Reizung; Sensibilisierung; Narkose", "CMR; STOT; Aspiration", "Gewässergefährdend"],
        "prob_scale_title": "WAHRSCHEINLICHKEITSSKALA:",
        "prob_scale": [("1 - Unwahrscheinlich", "Selten, wirksamer Schutz"), ("2 - Möglich", "Gelegentlich, teilweiser Schutz"), ("3 - Wahrscheinlich", "Regelmäßig, unzureichend"), ("4 - Sehr wahrscheinlich", "Dauerhaft, kein Schutz")],
        "sev_scale_title": "SCHWEREGRADSKALA:",
        "sev_scale": [("1 - Vernachlässigbar", "Leicht, reversibel"), ("2 - Gering", "Reversible Schädigung"), ("3 - Mittel", "Schwer, potenziell irreversibel"), ("4 - Schwer", "Tödlich/bleibend, CMR")],
        "db_headers": ["Nr.", "Produktkategorie", "Handelsname", "SDB-Sprache", "SDB-Version", "Ausgabedatum", "Überarbeitungsdatum", "Hersteller", "Adresse", "Telefon", "E-Mail", "Notruf", "UFI-Code", "Produktform", "Verwendung", "Kategorie", "Stoff/Gemisch", "Bestandteil 1 - Name", "Bestandteil 1 - CAS", "Bestandteil 1 - EG", "Bestandteil 1 - Konz.%", "Bestandteil 1 - CLP", "Bestandteil 2 - Name", "Bestandteil 2 - CAS", "Bestandteil 2 - EG", "Bestandteil 2 - Konz.%", "Bestandteil 2 - CLP", "Bestandteil 3 - Name", "Bestandteil 3 - CAS", "Bestandteil 3 - Konz.%", "Bestandteil 3 - CLP", "CLP-Einstufung (Gemisch)", "GHS-Piktogramme", "Signalwort", "H-Sätze", "P-Sätze", "EUH-Sätze", "SVHC", "PBT/vPvB", "Aggregatzustand", "Farbe", "Geruch", "Schmelzpunkt (°C)", "Siedepunkt (°C)", "Flammpunkt (°C)", "Selbstentzündung (°C)", "Dichte (g/cm³)", "Wasserlöslichkeit", "pH", "Dampfdruck", "AGW (mg/m³)", "KZE (mg/m³)", "MAK (mg/m³)", "DNEL Inhalation", "DNEL dermal", "BOELV (EU) mg/m³", "Atemschutz", "Handschutz", "Augenschutz", "Hautschutz", "Technische Maßnahmen", "Löschmittel geeignet", "Löschmittel ungeeignet", "Zersetzungsprodukte", "Feuerwehr-PSA", "Orale LD50", "Dermale LD50", "Inhal. LC50", "Hautreizung", "Augenreizung", "Sensibilisierung", "CMR-Wirkungen", "UN-Nr.", "Versandbezeichnung", "ADR-Klasse", "Verpackungsgruppe", "Meeresschadstoff", "EAK-Code", "Entsorgung", "Verwendungsort", "Jahresmenge", "Häufigkeit", "Expositionsweg", "Exponierte MA", "Bemerkungen"],
        "risk_headers": ["Nr.", "Handelsname", "Hauptbestandteil", "CLP-Einstufung", "H-Sätze", "P-Sätze", "Expositionsweg", "Häufigkeit", "Dauer", "Betroffene Körperteile", "Schutz vorhanden", "PSA-Spezifikation", "Wahrscheinlichkeit (1-4)", "Schweregrad (1-4)", "Risiko (WxS)", "Risikostufe", "Maßnahme", "BEM nötig", "Expositionsverzeichnis Pflicht", "Frist", "Verantwortlich", "Wahrsch. danach", "Schwere danach", "Restrisiko", "Restrisikostufe", "Beurteiler", "Datum", "Überprüfung", "Bemerkungen"],
        "exp_headers": ["Nr.", "Mitarbeiter", "Geburtsort/-datum", "Muttername", "Beruf", "Arbeitsplatz", "Gefahrstoff", "CAS-Nr.", "Expositionsweg", "Tägliche Exp. (h)", "Wöchentliche Exp. (h)", "Jährliche Exp. (h)", "Gemessene Konz. (mg/m³)", "AGW/KZE (mg/m³)", "PSA", "Arb.med. Untersuchung", "Beginn", "Bemerkungen"],
        "exp_note": "Expositionsverzeichnis - Vom Arbeitgeber pro Mitarbeiter auszufüllen!",
        "action_headers": ["Nr.", "Gefahrstoff", "Risikostufe", "Maßnahme", "Verantwortlich", "Frist", "Status", "Abschluss", "Bemerkungen"],
        "use_location": "Produktion", "company_fills": "Vom Unternehmen!", "employer": "Arbeitgeber", "in_progress": "In Bearbeitung",
    },
}

def get_lang(lang_code):
    if lang_code in LANG_STRINGS:
        return LANG_STRINGS[lang_code]
    return LANG_STRINGS["en"]

# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.header("⚙️ Beállítások")
    api_key = ""
    try:
        api_key = st.secrets.get("OPENAI_API_KEY", "")
        if not api_key:
            api_key = st.secrets.get("MY_API_KEY", "")
    except:
        pass
    if not api_key:
        api_key = st.text_input("🔑 OpenAI API kulcs", type="password")
    else:
        st.success("✅ API kulcs betöltve")

    st.divider()
    st.subheader("🌍 Kimeneti nyelv")
    output_lang_label = st.selectbox("📝 Az Excel kimeneti nyelve", list(LANGUAGES.keys()), index=0,
                                     help="Fejlécek, útmutató, AI kimenet ezen a nyelven készül")
    output_lang = LANGUAGES[output_lang_label]

    st.divider()
    evaluator_name = st.text_input("👤 Értékelő neve", value="")
    eval_date = st.date_input("📅 Értékelés dátuma", value=datetime.now())
    review_date = st.date_input("📅 Felülvizsgálat", value=datetime(datetime.now().year + 1, datetime.now().month, datetime.now().day))
    deadline_date = st.date_input("📅 Intézkedés határideje", value=datetime(datetime.now().year, 6, 30))

# ============================================================
# FIZETÉSI FAL – az alkalmazás indulása előtt
# ============================================================
show_payment_wall()

# ============================================================
# SYSTEM PROMPTS
# ============================================================

SYSTEM_PROMPT_EXTRACT = """Te egy veszélyes anyagok szakértője vagy, aki biztonsági adatlapokból (SDS/MSDS) nyer ki strukturált adatokat.
FELADATOD: Olvasd el a biztonsági adatlapot és válaszolj KIZÁRÓLAG érvényes JSON formátumban.
NYELVI SZABÁLYOK:
- Bármilyen nyelvű SDS-t elfogadsz → kimenet a MEGADOTT CÉLNYELVEN
- A célnyelvet a user message tartalmazza
- Kémiai neveknél a célnyelv szakkifejezéseit használd
KRITIKUS: A H és P mondatoknál a KÓDOT ÉS a TELJES SZÖVEGET add meg A MEGADOTT CÉLNYELVEN!
Példa (magyar): "H225 (Fokozottan tűzveszélyes folyadék és gőz); H319 (Súlyos szemirritációt okoz)"
Példa (English): "H225 (Highly flammable liquid and vapour); H319 (Causes serious eye irritation)"
Példa (Deutsch): "H225 (Flüssigkeit und Dampf leicht entzündbar); H319 (Verursacht schwere Augenreizung)"
FORDÍTSD le a célnyelvre!
Az összetevők CLP osztályzásánál is add meg a H mondatot kifejtve.
JSON SÉMA:
{"product_name":"","product_category":"","sds_language":"","sds_version":"","sds_date":"","sds_revision_date":"","manufacturer":"","manufacturer_address":"","manufacturer_phone":"","manufacturer_email":"","emergency_phone":"","ufi_code":"","product_form":"","intended_use":"","use_category":"","substance_or_mixture":"","comp1_name":"","comp1_cas":"","comp1_ec":"","comp1_conc":"","comp1_clp":"","comp2_name":"","comp2_cas":"","comp2_ec":"","comp2_conc":"","comp2_clp":"","comp3_name":"","comp3_cas":"","comp3_conc":"","comp3_clp":"","clp_classification":"","ghs_pictograms":"","signal_word":"","h_statements":"","p_statements":"","euh_statements":"","svhc":"","pbt_vpvb":"","physical_state":"","color":"","odor":"","melting_point":"","boiling_point":"","flash_point":"","autoignition_temp":"","density":"","water_solubility":"","ph":"","vapor_pressure":"","ak_value":"","ck_value":"","mk_value":"","dnel_inhalation":"","dnel_dermal":"","boelv":"","respiratory_protection":"","hand_protection":"","eye_protection":"","skin_protection":"","engineering_controls":"","suitable_extinguishing":"","unsuitable_extinguishing":"","hazardous_decomposition":"","firefighter_ppe":"","ld50_oral":"","ld50_dermal":"","lc50_inhalation":"","skin_irritation":"","eye_irritation":"","sensitization":"","cmr_effects":"","un_number":"","shipping_name":"","adr_class":"","packing_group":"","marine_pollutant":"","ewc_code":"","disposal_method":"","exposure_routes":""}
FONTOS: Válaszolj KIZÁRÓLAG a fenti JSON-nal!"""

SYSTEM_PROMPT_RISK = """Te egy munkavédelmi kockázatértékelési szakértő vagy.
4x4 mátrix: V(1-4) × S(1-4). Szintek: 1-2 Alacsony, 3-4 Közepes, 5-9 Magas, 10-16 Elfogadhatatlan.
VÉDŐESZKÖZ SPECIFIKÁCIÓ: kesztyű anyag+vastagság+áttörési idő+EN szabvány; szűrő típus; szemvédő típus; bőrvédelem EN szab.
BEM: ólom/CMR → kötelező. Expozíciós nyilv.: CLP veszélyes → kötelező.
A CÉLNYELVEN válaszolj (a user message tartalmazza)!
JSON: {"main_hazardous_component":"","exposure_mode":"","exposure_frequency":"","exposure_duration":"","affected_body_parts":"","protection_present":"","ppe_specification":"","probability":2,"severity":3,"risk_score":6,"risk_level":"","required_action":"","bem_required":"","exposure_registry_required":"","post_action_probability":1,"post_action_severity":3,"residual_risk":3,"residual_risk_level":""}
FONTOS: Válaszolj KIZÁRÓLAG JSON-nal!"""

# ============================================================
# PDF + GPT
# ============================================================

def extract_text_from_pdf(pdf_file):
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(pdf_file)
        return "\n".join(p.extract_text() or "" for p in reader.pages)
    except Exception as e:
        return f"HIBA: {e}"

def call_gpt(api_key, system, user_msg):
    from openai import OpenAI
    client = OpenAI(api_key=api_key)
    try:
        resp = client.chat.completions.create(
            model="gpt-4o", messages=[{"role":"system","content":system},{"role":"user","content":user_msg}],
            temperature=0.1, max_tokens=4000, response_format={"type":"json_object"})
        r = json.loads(resp.choices[0].message.content)
        r['_tokens'] = resp.usage.total_tokens if resp.usage else 0
        return r
    except Exception as e:
        return {'_error': str(e)}

def process_single_sds(pdf_file, api_key, target_lang="hu"):
    lang_name = LANG_NAMES.get(target_lang, target_lang)
    pdf_text = extract_text_from_pdf(pdf_file)
    if pdf_text.startswith("HIBA") or len(pdf_text.strip()) < 100:
        return {'_source_file': pdf_file.name, '_status': '❌ PDF hiba'}, {}
    if len(pdf_text) > 25000:
        pdf_text = pdf_text[:25000] + "\n[...]"
    sds = call_gpt(api_key, SYSTEM_PROMPT_EXTRACT,
                   f"CÉLNYELV: {lang_name}\nA kimenet {lang_name} nyelven legyen!\n\n{pdf_text}")
    if '_error' in sds:
        sds['_source_file'] = pdf_file.name; sds['_status'] = f"❌ {sds['_error']}"
        return sds, {}
    sds['_source_file'] = pdf_file.name; sds['_status'] = '✅'
    ri = json.dumps({k:v for k,v in sds.items() if not k.startswith('_')}, ensure_ascii=False, indent=1)
    risk = call_gpt(api_key, SYSTEM_PROMPT_RISK,
                    f"CÉLNYELV: {lang_name}\nKészíts kockázatértékelést {lang_name} nyelven:\n\n{ri}")
    if '_error' in risk: risk = {}
    return sds, risk

# ============================================================
# EXCEL GENERÁLÁS
# ============================================================

def generate_full_excel(results, risk_results, evaluator, eval_date, review_date, deadline_date, lang_code="hu"):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = openpyxl.Workbook()
    L = get_lang(lang_code)

    DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ORANGE = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    WHITE_FONT = Font(color="FFFFFF", bold=True, size=10)
    BOLD = Font(bold=True, size=10)
    NORMAL = Font(size=9)
    WRAP = Alignment(wrap_text=True, vertical='top')
    CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)
    THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 1. ÚTMUTATÓ
    ws1 = wb.active; ws1.title = L["sheet_names"][0]; ws1.sheet_properties.tabColor = "1F4E79"
    guide = [
        (L["main_title"], True), ("", False), (L["prepared_by"], False),
        (f"{L['prep_date']}: {datetime.now().strftime('%Y.%m.%d.')}", False),
        (f"{L['processed_count']}: {len(results)} db", False), ("", False),
        (L["legal_bg"], True),
        ("• 1993. évi XCIII. tv. (Mvt.) - 54.§, 63/A.§", False),
        ("• 2000. évi XXV. tv. (Kbtv.)", False),
        ("• 5/2020. (II. 6.) ITM rendelet", False),
        ("• 25/2000. (IX. 30.) EüM-SzCsM rendelet", False),
        ("• 1272/2008/EK (CLP)", False), ("• 1907/2006/EK (REACH)", False),
        ("• (EU) 2020/878", False), ("", False),
        (L["sheets_content"], True),
    ]
    for i, desc in enumerate(L["sheet_desc"]):
        guide.append((f"{i+1}. {desc}", False))
    guide += [("", False), (L["markings"], False), (L["empty_cells"], False)]
    for ri, (txt, bold) in enumerate(guide, 1):
        c = ws1.cell(row=ri, column=1, value=txt)
        c.font = Font(bold=bold, size=14 if ri==1 else 10, color="FFFFFF" if ri==1 else "000000")
        if ri == 1: c.fill = DARK_BLUE
    ws1.column_dimensions['A'].width = 120

    # 2. SEGÉDTÁBLÁK
    ws6 = wb.create_sheet(L["sheet_names"][1]); ws6.sheet_properties.tabColor = "7030A0"
    c = ws6.cell(row=1, column=1, value=L["risk_matrix_title"]); c.fill = DARK_BLUE; c.font = WHITE_FONT
    ws6.merge_cells('A1:E1')
    for ci, h in enumerate([""] + L["severity"], 1):
        c = ws6.cell(row=3, column=ci, value=h); c.font = BOLD; c.alignment = CENTER; c.border = THIN
    matrix = [[L["probability"][0],4,8,12,16],[L["probability"][1],3,6,9,12],[L["probability"][2],2,4,6,8],[L["probability"][3],1,2,3,4]]
    for ri, row in enumerate(matrix, 4):
        for ci, val in enumerate(row, 1):
            c = ws6.cell(row=ri, column=ci, value=val); c.border = THIN; c.alignment = CENTER
            if isinstance(val, int):
                if val <= 2: c.fill = GREEN
                elif val <= 4: c.fill = YELLOW
                elif val <= 9: c.fill = ORANGE
                else: c.fill = RED_FILL
    rl_data = [(9, L["risk_levels_title"], None),(10, L["risk_levels"][0], GREEN),(11, L["risk_levels"][1], YELLOW),(12, L["risk_levels"][2], ORANGE),(13, L["risk_levels"][3], RED_FILL)]
    for ri, txt, fill in rl_data:
        c = ws6.cell(row=ri, column=1, value=txt); c.font = BOLD if ri==9 else NORMAL
        if fill: c.fill = fill
    ws6.cell(row=15, column=1, value=L["ghs_title"]).font = BOLD
    ghs_codes = ["GHS01","GHS02","GHS03","GHS04","GHS05","GHS06","GHS07","GHS08","GHS09"]
    for ri, (gc, gs, gd) in enumerate(zip(ghs_codes, L["ghs_symbols"], L["ghs_desc"]), 16):
        ws6.cell(row=ri, column=1, value=gc).font = BOLD; ws6.cell(row=ri, column=2, value=gs); ws6.cell(row=ri, column=3, value=gd)
    ws6.cell(row=26, column=1, value=L["prob_scale_title"]).font = BOLD
    for ri, (lev, desc) in enumerate(L["prob_scale"], 27):
        ws6.cell(row=ri, column=1, value=lev).font = BOLD; ws6.cell(row=ri, column=2, value=desc)
    ws6.cell(row=32, column=1, value=L["sev_scale_title"]).font = BOLD
    for ri, (lev, desc) in enumerate(L["sev_scale"], 33):
        ws6.cell(row=ri, column=1, value=lev).font = BOLD; ws6.cell(row=ri, column=2, value=desc)
    ws6.column_dimensions['A'].width = 40; ws6.column_dimensions['B'].width = 30; ws6.column_dimensions['C'].width = 50

    # 3. ADATBÁZIS
    ws2 = wb.create_sheet(L["sheet_names"][2]); ws2.sheet_properties.tabColor = "00B050"
    db_h = L["db_headers"]
    db_keys = [None,"product_category","product_name","sds_language","sds_version","sds_date","sds_revision_date","manufacturer","manufacturer_address","manufacturer_phone","manufacturer_email","emergency_phone","ufi_code","product_form","intended_use","use_category","substance_or_mixture","comp1_name","comp1_cas","comp1_ec","comp1_conc","comp1_clp","comp2_name","comp2_cas","comp2_ec","comp2_conc","comp2_clp","comp3_name","comp3_cas","comp3_conc","comp3_clp","clp_classification","ghs_pictograms","signal_word","h_statements","p_statements","euh_statements","svhc","pbt_vpvb","physical_state","color","odor","melting_point","boiling_point","flash_point","autoignition_temp","density","water_solubility","ph","vapor_pressure","ak_value","ck_value","mk_value","dnel_inhalation","dnel_dermal","boelv","respiratory_protection","hand_protection","eye_protection","skin_protection","engineering_controls","suitable_extinguishing","unsuitable_extinguishing","hazardous_decomposition","firefighter_ppe","ld50_oral","ld50_dermal","lc50_inhalation","skin_irritation","eye_irritation","sensitization","cmr_effects","un_number","shipping_name","adr_class","packing_group","marine_pollutant","ewc_code","disposal_method","_loc","_qty","_freq","exposure_routes","_workers","_notes"]
    for ci, h in enumerate(db_h, 1):
        c = ws2.cell(row=1, column=ci, value=h); c.fill = DARK_BLUE; c.font = WHITE_FONT; c.alignment = CENTER; c.border = THIN
    for ri, data in enumerate(results, 2):
        for ci, key in enumerate(db_keys, 1):
            if key is None: val = ri-1
            elif key == "_loc": val = L["use_location"]
            elif key in ("_qty","_freq","_workers"): val = L["company_fills"]
            elif key == "_notes": val = ""
            else: val = data.get(key, '') or ''
            c = ws2.cell(row=ri, column=ci, value=str(val)); c.border = THIN; c.alignment = WRAP; c.font = NORMAL
    for ci in range(1, len(db_h)+1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 60 if ci in (35,36) else 20 if ci > 5 else 12
    ws2.column_dimensions['C'].width = 30
    ws2.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(db_h))}1"; ws2.freeze_panes = 'D2'

    # 4. KOCKÁZATÉRTÉKELÉS
    ws3 = wb.create_sheet(L["sheet_names"][3]); ws3.sheet_properties.tabColor = "FF0000"
    rh = L["risk_headers"]
    for ci, h in enumerate(rh, 1):
        c = ws3.cell(row=1, column=ci, value=h); c.fill = DARK_BLUE; c.font = WHITE_FONT; c.alignment = CENTER; c.border = THIN
    for ri, (sds, risk) in enumerate(zip(results, risk_results), 2):
        if not risk: risk = {}
        rd = [ri-1, sds.get('product_name',''), risk.get('main_hazardous_component',sds.get('comp1_name','')),
              sds.get('clp_classification',''), sds.get('h_statements',''), sds.get('p_statements',''),
              risk.get('exposure_mode',sds.get('exposure_routes','')), risk.get('exposure_frequency',''),
              risk.get('exposure_duration',''), risk.get('affected_body_parts',''),
              risk.get('protection_present',''), risk.get('ppe_specification',''),
              risk.get('probability',''), risk.get('severity',''), risk.get('risk_score',''),
              risk.get('risk_level',''), risk.get('required_action',''),
              risk.get('bem_required',''), risk.get('exposure_registry_required',''),
              deadline_date.strftime('%Y.%m.%d'), L["employer"],
              risk.get('post_action_probability',''), risk.get('post_action_severity',''),
              risk.get('residual_risk',''), risk.get('residual_risk_level',''),
              evaluator, eval_date.strftime('%Y.%m.%d'), review_date.strftime('%Y.%m.%d'), '']
        for ci, val in enumerate(rd, 1):
            c = ws3.cell(row=ri, column=ci, value=str(val) if val else ''); c.border = THIN; c.alignment = WRAP; c.font = NORMAL
            if ci in (16, 25):
                lev = str(val).lower()
                if any(x in lev for x in ['alacsony','zöld','elfogadhat','acceptable','akzeptabel','green','grün']): c.fill = GREEN
                elif any(x in lev for x in ['közepes','sárga','tolerál','tolerable','yellow','gelb']): c.fill = YELLOW
                elif any(x in lev for x in ['magas','narancs','jelentős','significant','orange','erheblich']): c.fill = ORANGE
                elif any(x in lev for x in ['elfogadhatatlan','piros','unacceptable','red','inakzeptabel','rot']): c.fill = RED_FILL
    for ci in range(1, len(rh)+1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 60 if ci==12 else 50 if ci in (5,6) else 25 if ci > 3 else 12
    ws3.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(rh))}1"; ws3.freeze_panes = 'C2'

    # 5. EXPOZÍCIÓS NYILVÁNTARTÁS
    ws4 = wb.create_sheet(L["sheet_names"][4]); ws4.sheet_properties.tabColor = "FFC000"
    eh = L["exp_headers"]
    for ci, h in enumerate(eh, 1):
        c = ws4.cell(row=1, column=ci, value=h); c.fill = DARK_BLUE; c.font = WHITE_FONT; c.alignment = CENTER; c.border = THIN
    c = ws4.cell(row=2, column=1, value=L["exp_note"]); c.font = Font(bold=True, italic=True, size=10, color="FF0000")
    ws4.merge_cells(f'A2:{openpyxl.utils.get_column_letter(len(eh))}2')
    for ci in range(1, len(eh)+1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 20

    # 6. INTÉZKEDÉSI TERV
    ws5 = wb.create_sheet(L["sheet_names"][5]); ws5.sheet_properties.tabColor = "FF6600"
    ah = L["action_headers"]
    for ci, h in enumerate(ah, 1):
        c = ws5.cell(row=1, column=ci, value=h); c.fill = DARK_BLUE; c.font = WHITE_FONT; c.alignment = CENTER; c.border = THIN
    ar = 2
    for sds, risk in zip(results, risk_results):
        if risk:
            try: score = int(risk.get('risk_score', 0))
            except: score = 0
            if score >= 3:
                for ci, val in enumerate([ar-1, sds.get('product_name',''), risk.get('risk_level',''),
                    risk.get('required_action',''), L["employer"], deadline_date.strftime('%Y.%m.%d'),
                    L["in_progress"], '', ''], 1):
                    c = ws5.cell(row=ar, column=ci, value=str(val) if val else ''); c.border = THIN; c.alignment = WRAP
                ar += 1
    for ci in range(1, len(ah)+1):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 50 if ci==4 else 20

    wb.move_sheet(L["sheet_names"][1], offset=-2)

    output = BytesIO(); wb.save(output); output.seek(0)
    return output.getvalue()

# ============================================================
# FŐ FELÜLET
# ============================================================
st.title("🧪 SDS → Excel AI Feldolgozó v3.2")
st.markdown(f"**{output_lang_label}** | 6 munkalap | H/P kifejtés | Védőeszköz spec. | Kockázatértékelés")

# Fizetési státusz megjelenítése
if st.session_state.get("barion_paid"):
    st.success("✅ Barion fizetés sikeres – az alkalmazás használatra kész!")
elif st.session_state.get("user_subscribed"):
    st.success("✅ Stripe előfizetés aktív – az alkalmazás használatra kész!")

uploaded = st.file_uploader("📤 PDF biztonsági adatlapok", type=["pdf"], accept_multiple_files=True)

if uploaded:
    c1, c2, c3 = st.columns(3)
    c1.metric("📄 Fájlok", len(uploaded)); c2.metric("⏱️ Idő", f"~{len(uploaded)*30}s"); c3.metric("💰 Költség", f"~${len(uploaded)*0.30:.2f}")

    if not api_key:
        st.error("⚠️ Add meg az OpenAI API kulcsot!")
    elif not evaluator_name:
        st.warning("⚠️ Add meg az értékelő nevét!")
    elif st.button("🚀 FELDOLGOZÁS INDÍTÁSA", type="primary", use_container_width=True):
        prog = st.progress(0); status = st.empty(); log = st.container()
        all_r, all_k = [], []
        for i, pdf in enumerate(uploaded):
            prog.progress(i/len(uploaded), f"📄 {pdf.name} ({i+1}/{len(uploaded)})")
            status.info(f"🔄 **{pdf.name}** – feldolgozás {output_lang_label} nyelven...")
            sds, risk = process_single_sds(pdf, api_key, output_lang)
            all_r.append(sds); all_k.append(risk)
            with log: st.text(f"  {sds.get('_status','?')} {sds.get('product_name','?')} | {risk.get('risk_level','—') if risk else '—'}")
            time.sleep(0.3)
        prog.progress(1.0, "✅ Kész!"); status.success(f"✅ {len(all_r)} SDS feldolgozva ({output_lang_label})")
        st.session_state.results = all_r; st.session_state.risk_results = all_k; st.session_state.processing_done = True

if st.session_state.processing_done and st.session_state.results:
    results = st.session_state.results
    risks = st.session_state.risk_results

    st.divider()
    st.header("📊 Eredmények")

    tbl = [{
        "Státusz": s.get("_status", "?"),
        "Termék": s.get("product_name", "—"),
        "Kockázat": r.get("risk_level", "—") if r else "—"
    } for s, r in zip(results, risks)]
    st.dataframe(tbl, use_container_width=True, hide_index=True)

    try:
        excel = generate_full_excel(
            results, risks,
            evaluator_name, eval_date, review_date, deadline_date,
            output_lang
        )
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        excel_filename = f"SDS_Database_{output_lang}_{ts}.xlsx"
        excel_path = os.path.join(os.getcwd(), excel_filename)

        # fájl mentése a lemezre – ezt használja majd a checker
        with open(excel_path, "wb") as f:
            f.write(excel)

        # letöltőgomb (ugyanaz a tartalom)
        st.download_button(
            f"📥 LETÖLTÉS ({output_lang_label})",
            data=excel,
            file_name=excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        # 1 kattintásos SDS verzióellenőrzés
        if st.button("🆕 SDS verziók ellenőrzése és frissítése", use_container_width=True):
            try:
                cmd = [sys.executable, "sds_version_checker.py", "--input", excel_path]
                subprocess.run(cmd, check=True)
                st.success(
                    "✅ SDS verzióellenőrzés lefutott, az újabb biztonsági adatlapok – ha elérhetők – letöltve."
                )
            except Exception as e:
                st.error(f"❌ Hiba a verzióellenőrzés futtatásakor: {e}")

    except Exception as e:
        st.error(f"❌ {e}")

st.divider()
st.caption("🧪 SDS AI v3.2 | 24 EU nyelv | 6 munkalap | H/P kifejtés | Védőeszköz spec. | Kockázatértékelés | © Safety Expert – safetyexpert.hu")
